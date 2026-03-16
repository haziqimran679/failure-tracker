import streamlit as st
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.gridspec import GridSpec
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
import os
import io
import shutil
from datetime import datetime, date
from pathlib import Path

# ── Config ──────────────────────────────────────────────────────────────────
EXCEL_PATH = Path(r"T:\SCOF PE\Yield\Diva\Daily Update DIVA 50 module\failure_log.xlsx")
PHOTOS_DIR = Path("photos")
PHOTOS_DIR.mkdir(exist_ok=True)

PROCESSES = [
    "Pretest",
    "Adjustment",
    "Minirel",
    "Verification",
]
FAILURE_TYPES = [
    "DigitalInitializeCheck",
    "ABUS",
    "KDMI",
    "Adjust_ReferenceFrequency",
    "Adjust_IFRangingGain",
    "Adjust_SourceInsertionLossRatio",
    "Adjust_Compression",
    "Adjust_WaveshaperIDAC",
    "Adjust_PreVirtualBridge",
    "Adjust_StepAttenuatorPort1",
    "Adjust_RchGain",
    "Adjust_StepAttenuatorOtherPorts",
    "Adjust_SourceLevelDAC",
    "Adjust_VirtualBridge",
    "Adjust_SpectrumAnalyzer_Gain",
    "Adjust_IFFlatness",
    "IFFlatness",
    "PowerLinearityWithPowerSweep",
    "PowerLevelTransient",
    "TraceNoiseThrough",
    "Harmonics",
    "UnratioedPowerMeasurement",
    "SourceOnOffRatio",
    "UncorrectedPortChar",
    "CompressionRch",
    "PulseProfile_PreCheck",
    "OutputLevelRange",
    "RampSweep",
    "NoiseFloor",
    "Crosstalk",
    "IFRanging",
    "TraceNoiseOpen",
    "CompressionOpen",
    "OutputAccuracyOddPorts",
    "LOLeakageOddPorts",
    "OutputAccuracyEvenPorts",
    "LOLeakageEvenPorts",
    "Compression12",
    "FinalCheckProductID",
    "DDR3_Reset",
    "DDR3_Function",
    "Adjust_WriteNominalFiles",
    "WritePCASerialNumber",
    "LED",
    "LoopCountIncrement",
    "UnratioedPowerMeasurementOpen",
    "NoiseFloorOpen",
    "DDR3",
    "SweepWithDDR3Tests",
    "LoopCountClose",
]
COLUMNS = [
    "Date", "Model", "Serial Number", "Process",
    "Failure Type", "Description", "Remark", "Photo Path"
]

# ── Excel helpers ────────────────────────────────────────────────────────────

def _header_style(ws):
    """Apply styled header row to the Failure_Log sheet."""
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    col_widths = [14, 12, 16, 16, 18, 35, 20, 30]
    for col_idx, (col_name, width) in enumerate(zip(COLUMNS, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def init_excel():
    """Create Excel file with Failure_Log sheet if it does not exist."""
    if EXCEL_PATH.exists():
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Failure_Log"
    _header_style(ws)
    wb.save(EXCEL_PATH)


def load_records() -> pd.DataFrame:
    """Load all records from Excel into a DataFrame."""
    init_excel()
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name="Failure_Log", dtype=str)
        df.fillna("", inplace=True)
        if "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
        return df
    except Exception:
        return pd.DataFrame(columns=COLUMNS)


def append_record(row: dict):
    """Append a single record row to Excel, preserving existing data."""
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Failure_Log"]
    next_row = ws.max_row + 1

    # Alternate row fill for readability
    alt_fill = PatternFill("solid", fgColor="EBF3FB") if next_row % 2 == 0 else None
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=row.get(col, ""))
        cell.alignment = Alignment(vertical="center", wrap_text=(col in ("Description", "Remark")))
        cell.border = border
        if alt_fill:
            cell.fill = alt_fill

    # Highlight duplicate serial numbers in orange
    sn = row.get("Serial Number", "")
    if sn:
        sn_col = COLUMNS.index("Serial Number") + 1
        dup_fill = PatternFill("solid", fgColor="FFF0CC")
        count = 0
        for r in range(2, next_row):
            if ws.cell(row=r, column=sn_col).value == sn:
                count += 1
                ws.cell(row=r, column=sn_col).fill = dup_fill
        if count > 0:
            ws.cell(row=next_row, column=sn_col).fill = PatternFill("solid", fgColor="FFD966")

    ws.row_dimensions[next_row].height = 18
    wb.save(EXCEL_PATH)


def delete_record(row_index: int):
    """Delete a record by its 0-based DataFrame index (Excel row = index + 2)."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Failure_Log"]
    excel_row = row_index + 2          # header is row 1
    ws.delete_rows(excel_row)
    wb.save(EXCEL_PATH)


# ── Pareto helpers ───────────────────────────────────────────────────────────

def compute_pareto(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "Failure Type" not in df.columns:
        return pd.DataFrame()
    counts = (
        df["Failure Type"]
        .value_counts()
        .reset_index()
    )
    counts.columns = ["Failure Type", "Count"]
    total = counts["Count"].sum()
    counts["Percentage (%)"] = (counts["Count"] / total * 100).round(1)
    counts["Cumulative (%)"] = counts["Percentage (%)"].cumsum().round(1)
    return counts


def plot_pareto(pareto_df: pd.DataFrame, title: str = "Pareto Chart") -> plt.Figure:
    fig = plt.figure(figsize=(10, 5))
    gs = GridSpec(1, 1, figure=fig)
    ax1 = fig.add_subplot(gs[0])
    ax2 = ax1.twinx()

    labels = pareto_df["Failure Type"].tolist()
    counts = pareto_df["Count"].tolist()
    cum_pct = pareto_df["Cumulative (%)"].tolist()
    x = range(len(labels))

    # Bars
    bars = ax1.bar(x, counts, color="#185fa5", width=0.55, zorder=3, label="Count")
    ax1.set_xticks(list(x))
    ax1.set_xticklabels(labels, rotation=20, ha="right", fontsize=9)
    ax1.set_ylabel("Failure Count", fontsize=10)
    ax1.yaxis.set_major_locator(mticker.MaxNLocator(integer=True))
    ax1.grid(axis="y", linestyle="--", alpha=0.4, zorder=0)
    ax1.set_axisbelow(True)

    # Bar value labels
    for bar, v in zip(bars, counts):
        ax1.text(bar.get_x() + bar.get_width() / 2,
                 bar.get_height() + 0.3, str(v),
                 ha="center", va="bottom", fontsize=8, color="#185fa5")

    # Cumulative line
    ax2.plot(list(x), cum_pct, color="#e24b4a", marker="o",
             markersize=5, linewidth=2, label="Cumulative %", zorder=4)
    ax2.axhline(80, color="#e24b4a", linestyle=":", linewidth=1, alpha=0.5)
    ax2.set_ylabel("Cumulative %", fontsize=10)
    ax2.set_ylim(0, 110)
    ax2.yaxis.set_major_formatter(mticker.PercentFormatter())

    # Legend
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2,
               loc="upper left", fontsize=9, framealpha=0.8)

    ax1.set_title(title, fontsize=12, fontweight="bold", pad=12)
    fig.tight_layout()
    return fig


def export_pareto_to_excel(pareto_df: pd.DataFrame,
                            start_date, end_date) -> bytes:
    """Build a standalone Excel report with the Pareto data + chart."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pareto_Report"

    # ── Title block ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "Failure Pareto Report"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    range_cell = ws["A2"]
    s = str(start_date) if start_date else "All"
    e = str(end_date) if end_date else "All"
    range_cell.value = f"Date Range: {s}  →  {e}    |    Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    range_cell.font = Font(size=10, color="444444")
    range_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16

    # ── Header row ────────────────────────────────────────────────────────────
    headers = ["#", "Failure Type", "Count", "Percentage (%)", "Cumulative (%)"]
    widths   = [6,   22,            10,      18,               18]
    hfill    = PatternFill("solid", fgColor="1F4E79")
    hfont    = Font(bold=True, color="FFFFFF", size=11)
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Data rows ─────────────────────────────────────────────────────────────
    bar_fill = PatternFill("solid", fgColor="D6E8F5")   # highlight top contributor
    for i, row in pareto_df.iterrows():
        excel_row = i + 5
        values = [i + 1, row["Failure Type"], row["Count"],
                  row["Percentage (%)"], row["Cumulative (%)"]]
        alt = PatternFill("solid", fgColor="EBF3FB") if excel_row % 2 == 0 else None
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=v)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            if alt:
                cell.fill = alt
        # Highlight top failure
        if i == 0:
            for col_idx in range(1, 6):
                ws.cell(row=excel_row, column=col_idx).fill = bar_fill
                ws.cell(row=excel_row, column=col_idx).font = Font(bold=True)

    # ── Bar chart ─────────────────────────────────────────────────────────────
    last_data_row = len(pareto_df) + 4
    bar = BarChart()
    bar.type = "col"
    bar.title = "Failure Count by Type"
    bar.y_axis.title = "Count"
    bar.x_axis.title = "Failure Type"
    bar.style = 10
    bar.width = 18
    bar.height = 12

    data = Reference(ws, min_col=3, min_row=4, max_row=last_data_row)
    cats = Reference(ws, min_col=2, min_row=5, max_row=last_data_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, "G4")

    # ── Line chart (cumulative) ───────────────────────────────────────────────
    line = LineChart()
    line.title = "Cumulative %"
    line.y_axis.title = "Cumulative %"
    line.style = 10
    line.width = 18
    line.height = 12

    cum_data = Reference(ws, min_col=5, min_row=4, max_row=last_data_row)
    line.add_data(cum_data, titles_from_data=True)
    line.set_categories(cats)
    ws.add_chart(line, "G22")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Streamlit UI ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Failure Tracking System",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── Custom CSS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
  [data-testid="stSidebar"] { background: #1e2a3a; }
  .block-container { padding-top: 1.5rem; }
  h1 { font-size: 1.5rem !important; }
  h2 { font-size: 1.2rem !important; }
  .metric-card {
    background: #EBF3FB; border-radius: 10px;
    padding: 14px 18px; text-align: center;
  }
  .metric-card .val { font-size: 2rem; font-weight: 600; color: #185fa5; }
  .metric-card .lbl { font-size: 0.78rem; color: #555; margin-top: 2px; }
  .dup-badge {
    background: #FFD966; color: #7a5500;
    padding: 1px 8px; border-radius: 20px;
    font-size: 0.75rem; font-weight: 600;
  }
  .new-badge {
    background: #C6EFCE; color: #276221;
    padding: 1px 8px; border-radius: 20px;
    font-size: 0.75rem; font-weight: 600;
  }
  div[data-testid="stForm"] { border: none; padding: 0; }
</style>
""", unsafe_allow_html=True)

# ── Sidebar Nav ──────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🔬 Failure Tracker")
    st.markdown("*Manufacturing QC System*")
    st.divider()
    page = st.radio(
        "Navigation",
        ["📋 Data Entry", "📊 Pareto Analysis", "📁 Failure Log"],
        label_visibility="collapsed"
    )
    st.divider()
    df_all = load_records()
    st.markdown(f"**Total Records:** {len(df_all)}")
    if not df_all.empty and "Failure Type" in df_all.columns:
        top = df_all["Failure Type"].value_counts().idxmax()
        st.markdown(f"**Top Failure:** {top}")
    st.divider()
    # Download full Excel log
    if EXCEL_PATH.exists():
        with open(EXCEL_PATH, "rb") as f:
            st.download_button(
                "⬇️ Download Full Excel Log",
                data=f.read(),
                file_name="failure_log.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

# ════════════════════════════════════════════════════════════════════════════
# PAGE 1 — DATA ENTRY
# ════════════════════════════════════════════════════════════════════════════
if page == "📋 Data Entry":
    st.title("📋 Record New Failure")

    with st.form("entry_form", clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            f_date    = st.date_input("Date *", value=date.today())
            f_model   = st.text_input("Model *", placeholder="e.g. ABC123")
            f_sn      = st.text_input("Serial Number *", placeholder="e.g. SN00981")
            f_process = st.selectbox("Process *", [""] + PROCESSES)
        with c2:
            f_type    = st.selectbox("Failure Type *", [""] + FAILURE_TYPES)
            f_remark  = st.text_input("Remark", placeholder="e.g. Reworked, Scrapped…")
            f_photo   = st.file_uploader("Photo (optional)", type=["jpg","jpeg","png","bmp","tiff"])

        f_desc = st.text_area("Description", placeholder="Describe the failure in detail…", height=80)

        submitted = st.form_submit_button("💾 Save Record", use_container_width=True, type="primary")

    if submitted:
        errors = []
        if not f_model.strip():  errors.append("Model")
        if not f_sn.strip():     errors.append("Serial Number")
        if not f_process:        errors.append("Process")
        if not f_type:           errors.append("Failure Type")

        if errors:
            st.error(f"Required fields missing: {', '.join(errors)}")
        else:
            photo_path = ""
            if f_photo:
                dest = PHOTOS_DIR / f_photo.name
                with open(dest, "wb") as out:
                    out.write(f_photo.read())
                photo_path = str(dest)

            append_record({
                "Date":          str(f_date),
                "Model":         f_model.strip(),
                "Serial Number": f_sn.strip(),
                "Process":       f_process,
                "Failure Type":  f_type,
                "Description":   f_desc.strip(),
                "Remark":        f_remark.strip(),
                "Photo Path":    photo_path,
            })
            st.success(f"✅ Record saved! Serial **{f_sn.strip()}** logged for {f_type}.")
            st.rerun()

    # Preview of recent entries
    df_all = load_records()
    if not df_all.empty:
        st.divider()
        st.subheader("Recent Entries")
        st.dataframe(
            df_all.tail(10).iloc[::-1].reset_index(drop=True),
            use_container_width=True, height=260
        )


# ════════════════════════════════════════════════════════════════════════════
# PAGE 2 — PARETO ANALYSIS
# ════════════════════════════════════════════════════════════════════════════
elif page == "📊 Pareto Analysis":
    st.title("📊 Pareto Analysis")

    df_all = load_records()

    # ── Date filter ──────────────────────────────────────────────────────────
    with st.container(border=True):
        fc1, fc2, fc3 = st.columns([2, 2, 1])
        with fc1:
            start_date = st.date_input("Start Date", value=None, key="d_start")
        with fc2:
            end_date   = st.date_input("End Date",   value=None, key="d_end")
        with fc3:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("Clear Filter", use_container_width=True):
                st.session_state.d_start = None
                st.session_state.d_end   = None
                st.rerun()

    # Apply filter
    df = df_all.copy()
    if not df.empty and "Date" in df.columns:
        if start_date:
            df = df[df["Date"] >= start_date]
        if end_date:
            df = df[df["Date"] <= end_date]

    pareto_df = compute_pareto(df)

    # ── Metrics ───────────────────────────────────────────────────────────────
    total = len(df)
    n_types = len(pareto_df)
    top_type  = pareto_df.iloc[0]["Failure Type"] if not pareto_df.empty else "—"
    top_pct   = pareto_df.iloc[0]["Percentage (%)"] if not pareto_df.empty else 0

    # Count how many types cover 80%
    pareto80 = 0
    if not pareto_df.empty:
        for _, row in pareto_df.iterrows():
            pareto80 += 1
            if row["Cumulative (%)"] >= 80:
                break

    m1, m2, m3, m4 = st.columns(4)
    for col, val, lbl, sub in [
        (m1, total,    "Total Failures",   "in selected range"),
        (m2, n_types,  "Failure Types",    "distinct types"),
        (m3, f"{top_pct}%", f"Top: {top_type}", "of total failures"),
        (m4, pareto80, "Types → 80% rule", "drive 80% of failures"),
    ]:
        col.markdown(
            f'<div class="metric-card"><div class="val">{val}</div>'
            f'<div class="lbl">{lbl}<br><span style="font-size:.7rem">{sub}</span></div></div>',
            unsafe_allow_html=True
        )

    st.markdown("")

    if pareto_df.empty:
        st.info("No data found for the selected date range.")
    else:
        # ── Chart ─────────────────────────────────────────────────────────────
        with st.container(border=True):
            st.subheader("Pareto Chart")
            title_str = "Pareto Chart"
            if start_date or end_date:
                s = str(start_date) if start_date else "All"
                e = str(end_date)   if end_date   else "All"
                title_str = f"Pareto Chart  ({s} → {e})"
            fig = plot_pareto(pareto_df, title=title_str)
            st.pyplot(fig, use_container_width=True)
            plt.close(fig)

        # ── Summary table ─────────────────────────────────────────────────────
        with st.container(border=True):
            st.subheader("Summary Table")
            display_df = pareto_df.copy()
            display_df.insert(0, "#", range(1, len(display_df) + 1))
            st.dataframe(display_df, use_container_width=True, hide_index=True)

        # ── Export button ─────────────────────────────────────────────────────
        excel_bytes = export_pareto_to_excel(pareto_df, start_date, end_date)
        fname = f"pareto_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button(
            "⬇️ Export Pareto Report to Excel",
            data=excel_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )


# ════════════════════════════════════════════════════════════════════════════
# PAGE 3 — FAILURE LOG
# ════════════════════════════════════════════════════════════════════════════
elif page == "📁 Failure Log":
    st.title("📁 Failure Log")

    df_all = load_records()

    if df_all.empty:
        st.info("No records yet. Use Data Entry to log failures.")
    else:
        # ── Search / filter bar ───────────────────────────────────────────────
        with st.container(border=True):
            sc1, sc2, sc3 = st.columns(3)
            search_sn   = sc1.text_input("Search Serial Number", placeholder="e.g. SN001")
            filter_type = sc2.selectbox("Filter by Failure Type", ["All"] + FAILURE_TYPES)
            filter_proc = sc3.selectbox("Filter by Process",      ["All"] + PROCESSES)

        df_view = df_all.copy()
        if search_sn:
            df_view = df_view[df_view["Serial Number"].str.contains(search_sn, case=False, na=False)]
        if filter_type != "All":
            df_view = df_view[df_view["Failure Type"] == filter_type]
        if filter_proc != "All":
            df_view = df_view[df_view["Process"] == filter_proc]

        st.markdown(f"**{len(df_view)}** records shown (of {len(df_all)} total)")

        # Mark duplicate serial numbers
        sn_counts = df_all["Serial Number"].value_counts()

        # Build display copy with badge column
        df_disp = df_view.copy().reset_index(drop=True)
        df_disp.insert(0, "Row", range(1, len(df_disp) + 1))
        df_disp["Dup?"] = df_disp["Serial Number"].apply(
            lambda sn: f"⚠️ ×{sn_counts[sn]}" if sn_counts.get(sn, 1) > 1 else "✅ New"
        )

        # Show styled dataframe
        st.dataframe(
            df_disp,
            use_container_width=True,
            height=420,
            column_config={
                "Row":            st.column_config.NumberColumn(width="small"),
                "Date":           st.column_config.DateColumn(format="YYYY-MM-DD"),
                "Photo Path":     st.column_config.TextColumn(width="medium"),
                "Dup?":           st.column_config.TextColumn("Repeat Unit?", width="small"),
            }
        )

        # ── Delete record ─────────────────────────────────────────────────────
        st.divider()
        with st.expander("🗑️ Delete a Record"):
            st.warning("Deleting is permanent and will update the Excel file immediately.")
            del_idx = st.number_input(
                "Row number to delete (from table above)",
                min_value=1, max_value=len(df_all), step=1
            )
            if st.button("Delete Record", type="primary"):
                delete_record(int(del_idx) - 1)
                st.success(f"Row {del_idx} deleted.")
                st.rerun()
